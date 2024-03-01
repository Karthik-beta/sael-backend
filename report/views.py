from django.shortcuts import render
from rest_framework import generics
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.pagination import PageNumberPagination
from datetime import datetime, timedelta
import openpyxl
from django.views.generic import View
from django.http import HttpResponse
from django.db.models import Q

from report import models
from report import serializers




class defaultPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 100


class productionReportList(generics.ListCreateAPIView):
    queryset = models.productionReport.objects.order_by('-id')
    serializer_class = serializers.productionReportSerializer
    pagination_class = defaultPagination



class ExportProductionReportExcelView(View):
    def get(self, request, *args, **kwargs):
        queryset = models.productionReport.objects.all()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Production Report"

        headers = ["Job ID", "Customer", "PO No", "Batch No", "Assigned Date", "Expected Date", "Company", "Plant", "Shopfloor", "Assembly Line", "Machine ID", "Product ID", "Date", "Shift", "Planned Hours", "Planned Production", "Remaining Hours", "Balance Production", "Manager"]
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)

        for row_num, record in enumerate(queryset, 2):
            ws.cell(row=row_num, column=1, value=record.job_id)
            ws.cell(row=row_num, column=2, value=record.customer)
            ws.cell(row=row_num, column=3, value=record.po_no)
            ws.cell(row=row_num, column=4, value=record.batch_no)
            ws.cell(row=row_num, column=5, value=record.assigned_date)
            ws.cell(row=row_num, column=6, value=record.expected_date)
            ws.cell(row=row_num, column=7, value=record.company)
            ws.cell(row=row_num, column=8, value=record.plant)
            ws.cell(row=row_num, column=9, value=record.shopfloor)
            ws.cell(row=row_num, column=10, value=record.assembly_line)
            ws.cell(row=row_num, column=11, value=record.machine_id)
            ws.cell(row=row_num, column=12, value=record.product_id)
            ws.cell(row=row_num, column=13, value=record.date)
            ws.cell(row=row_num, column=14, value=record.shift_a)
            ws.cell(row=row_num, column=15, value=record.shift_b)
            ws.cell(row=row_num, column=16, value=record.shift_c)
            ws.cell(row=row_num, column=17, value=record.planned_hours)
            ws.cell(row=row_num, column=18, value=record.planned_production)
            ws.cell(row=row_num, column=19, value=record.remaining_hours)
            ws.cell(row=row_num, column=20, value=record.balance_production)
            ws.cell(row=row_num, column=21, value=record.manager)

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = "attachment; filename=Production_Report.xlsx"
        wb.save(response)

        return response
    

# class ExportExcelMachineView(View):
#     def get(self, request, *args, **kwargs):
#         queryset = models.machineWiseData.objects.all()

#         wb = openpyxl.Workbook()
#         ws = wb.active
#         ws.title = "Machine Records"

#         headers = [
#             "Product ID", "Plant", "Shopfloor", "Assembly Line", "Machine ID",
#             "Product Target", "Time", "Date", "On Time", "Idle Time",
#             "Actual", "Target", "Performance", "Gap", "kW-h"
#         ]

#         for col_num, header in enumerate(headers, 1):
#             ws.cell(row=1, column=col_num, value=header)

#         for row_num, record in enumerate(queryset, 2):
#             ws.cell(row=row_num, column=1, value="AQUA 1000ml")
#             ws.cell(row=row_num, column=2, value=record.plant)
#             ws.cell(row=row_num, column=3, value=record.shopfloor)
#             ws.cell(row=row_num, column=4, value=record.assembly_line)
#             ws.cell(row=row_num, column=5, value=record.machine_id)
#             ws.cell(row=row_num, column=6, value=record.product_target)
#             ws.cell(row=row_num, column=7, value=record.time)
#             ws.cell(row=row_num, column=8, value=record.date)
#             ws.cell(row=row_num, column=9, value=record.on_time)
#             ws.cell(row=row_num, column=10, value=record.idle_time)
#             ws.cell(row=row_num, column=11, value=record.actual)
#             ws.cell(row=row_num, column=12, value=record.target)
#             ws.cell(row=row_num, column=13, value=record.performance)
#             ws.cell(row=row_num, column=14, value=record.performance)
#             ws.cell(row=row_num, column=15, value=record.current)

#         response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
#         response["Content-Disposition"] = "attachment; filename=machine_records.xlsx"
#         wb.save(response)

#         return response